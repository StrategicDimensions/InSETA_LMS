from collections import deque
import ast
import csv
import re
from openerp import http, _
from openerp.http import request
from openerp.tools import ustr
# from openerp.tools.misc import xlwt
# import xlsxwriter as xlwt
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# wb = load_workbook('filename.xlsx')
# wb = Workbook(write_only=True)
# .
# .
# .
# (make your edits)
# .
# .
# .
# wb.save('new_filename.xlsx')
try:
    import xlwt
except ImportError:
    xlwt = None

DEBUG = True

if DEBUG:
    import logging

    logger = logging.getLogger(__name__)


    def dbg(msg):
        logger.info(msg)
else:
    def dbg(msg):
        pass


def remove_ascii(text):
    return ''.join([i if ord(i) < 128 else ' ' for i in text])


class ReportExporter(http.Controller):
    @http.route('/web/pivot/check_xlwt', type='json', auth='none')
    def check_xlwt(self):
        return xlwt is not None

    @http.route(['/report_export/accreditation_analysis/<int:report_id>'], type='http', auth="user")
    def accreditation_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        accreds = request.env['seta.reports.accreditations'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        worksheet.write_merge(0, 0, 0, 7, _("accreditation report FROM %s TO %s") % (report.from_date, report.to_date),
                              header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, accred in enumerate(accreds):
            worksheet.write(i + 2, 0, accred.provider_accreditation_ref)
            worksheet.write(i + 2, 1, accred.name)
            worksheet.write(i + 2, 2, accred.phone)
            worksheet.write(i + 2, 3, accred.email)
            worksheet.write(i + 2, 4, accred.provider_register_date)
            worksheet.write(i + 2, 5, accred.provider_approval_date)
            worksheet.write(i + 2, 6, accred.is_extension_of_scope)
            worksheet.write(i + 2, 7, accred.is_existing_provider)

        # num_leads = len(leads)
        # worksheet.write_merge(num_leads + 2, num_leads + 2, 22, 23, "TOTAL", header_bold_blue)
        # worksheet.write(num_leads + 2, 24, xlwt.Formula("SUM($X$3:$X$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 25, xlwt.Formula("SUM($Y$3:$Y$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 26, xlwt.Formula("SUM($Z$3:$Z$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 27, xlwt.Formula("SUM($AA$3:$AA$%s)" % (num_leads + 2)), header_bold_yellow)
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=accreditation_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    @http.route(['/report_export/late_accreditation_analysis/<int:report_id>'], type='http', auth="user")
    def late_accreditation_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        accreds = request.env['seta.reports.late.accreditations'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        # report FROM %s TO %s") % (report.from_date, report.to_date)
        worksheet.write_merge(0, 0, 0, 8,
                              _("Provider Accreditation application assessed in 140 days per province at Provincial Level "
                                "province FROM %s TO %s") % (report.from_date, report.to_date),
                              header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, accred in enumerate(accreds):
            worksheet.write(i + 2, 0, accred.provider_name)
            worksheet.write(i + 2, 1, accred.provider_accreditation_ref)
            worksheet.write(i + 2, 2, accred.sdl)
            worksheet.write(i + 2, 3, accred.alt_sdl)
            worksheet.write(i + 2, 4, accred.state_id.name)
            worksheet.write(i + 2, 5, accred.application_date)
            worksheet.write(i + 2, 6, accred.days_to_assess)
            worksheet.write(i + 2, 7, accred.update_date)
            worksheet.write(i + 2, 8, accred.final_state)
            worksheet.write(i + 2, 9, accred.in_process)
            worksheet.write(i + 2, 10, accred.transaction_status)

        # num_leads = len(leads)
        # worksheet.write_merge(num_leads + 2, num_leads + 2, 22, 23, "TOTAL", header_bold_blue)
        # worksheet.write(num_leads + 2, 24, xlwt.Formula("SUM($X$3:$X$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 25, xlwt.Formula("SUM($Y$3:$Y$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 26, xlwt.Formula("SUM($Z$3:$Z$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 27, xlwt.Formula("SUM($AA$3:$AA$%s)" % (num_leads + 2)), header_bold_yellow)
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=late_accreditation_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    @http.route(['/report_export/assessment_analysis/<int:report_id>'], type='http', auth="user")
    def assessment_analysis(self, report_id, **kw):
        report = request.env['seta.reports'].search([('id', '=', report_id)])
        assessments = request.env['seta.reports.assessment'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        # remove duplicates
        duplicates_watchdog = []

        # /odoo_reports
        with open('/odoo_reports/assessment_analysis.csv', 'w') as csvfile:

            writer = csv.DictWriter(csvfile, fieldnames=headers, delimiter='~', quotechar='|')

            writer.writeheader()
            for assessment in assessments:
                # only selected choices  
                if report.assessment_state != assessment.assessment.state and report.assessment_state:
                    continue
                if report.qual_skill_assessment != assessment.qual_skill_assessment and report.qual_skill_assessment:
                    continue

                # check for duplicates 
                if assessment.assessment.name in duplicates_watchdog:
                    continue

                duplicates_watchdog.append(assessment.assessment.name)

                enrolled_count = 0

                if assessment.qual_skill_assessment == 'qual':
                    enrolled_count = len(list(set(assessment.assessment.learner_ids)))
                elif assessment.qual_skill_assessment == 'lp':
                    enrolled_count = len(list(set(assessment.assessment.learner_ids_for_lp)))
                elif assessment.qual_skill_assessment == 'skill':
                    enrolled_count = len(
                        list(set(assessment.assessment.learner_ids_for_skills)))  # included set to remove duplicates

                if not isinstance(assessment.batch_id.batch_name, bool):
                    assessment.batch_id.batch_name = assessment.batch_id.batch_name.replace(u"\xa0", u" ")
                    assessment.batch_id.batch_name = assessment.batch_id.batch_name.encode("utf-8")

                writer.writerow({'NAME': assessment.assessment.name,
                                 'provider': assessment.provider_id.name,
                                 'type': assessment.qual_skill_assessment,
                                 'batch': assessment.batch_id.batch_name,
                                 'state': assessment.assessment.state,
                                 'province': assessment.provider_province.name,
                                 'enrolled learners': enrolled_count,
                                 })

                if assessment.qual_skill_assessment == 'qual':
                    for learner in assessment.assessment.learner_ids:
                        rpl = False
                        foreign = False
                        lnr = learner.learner_id

                        if lnr.alternate_id_type:
                            foreign = True

                        # special case where both fields are filled
                        if lnr.learner_identification_id and lnr.national_id:
                            foreign = False

                        lnr.person_last_name = remove_ascii(lnr.person_last_name)
                        lnr.name = remove_ascii(lnr.name)

                        for quals in lnr.learner_qualification_ids:  # can make this a set
                            # if assessment.batch_id != quals.batch_id and str(assessment.batch_id) == str(quals.batch_id):
                            #    dbg(str(learner.display_name))
                            if assessment.batch_id == quals.batch_id:
                                for units in quals.learner_registration_line_ids:
                                    if units.is_rpl_learner:
                                        rpl = True

                                writer.writerow({'NAME': '',
                                                 'provider': '',
                                                 'type': '',
                                                 'batch': '',
                                                 'state': '',
                                                 'province': '',
                                                 'enrolled learners': '',
                                                 'learner ID': lnr.learner_identification_id,
                                                 'foreign ID': lnr.national_id if foreign else '',
                                                 'first name': lnr.name,
                                                 'last name': lnr.person_last_name,
                                                 'employed': lnr.socio_economic_status,
                                                 'rpl': rpl if rpl else '',
                                                 'achieved': quals.is_learner_achieved,
                                                 'qualification': learner.qual_learner_assessment_line_id.name,
                                                 'qualification id': quals.learner_qualification_parent_id.saqa_qual_id,
                                                 'learning programme id': '',
                                                 'skills programme id': '',
                                                 })

                if assessment.qual_skill_assessment == 'lp':
                    for learner in assessment.assessment.learner_ids_for_lp:

                        foreign = False
                        lnr = learner.learner_id

                        if lnr.alternate_id_type:
                            foreign = True

                        # special case where both fields are filled
                        if lnr.learner_identification_id and lnr.national_id:
                            foreign = False

                        lnr.person_last_name = remove_ascii(lnr.person_last_name)
                        lnr.name = remove_ascii(lnr.name)

                        for quals in lnr.learning_programme_ids:
                            if assessment.batch_id == quals.batch_id:
                                writer.writerow({'NAME': '',
                                                 'provider': '',
                                                 'type': '',
                                                 'batch': '',
                                                 'state': '',
                                                 'province': '',
                                                 'enrolled learners': '',
                                                 'learner ID': lnr.learner_identification_id,
                                                 'foreign ID': lnr.national_id if foreign else '',
                                                 'first name': lnr.name,
                                                 'last name': lnr.person_last_name,
                                                 'employed': lnr.socio_economic_status,
                                                 'achieved': quals.is_learner_achieved,
                                                 'qualification': learner.lp_learner_assessment_line_id.name,
                                                 'qualification id': quals.lp_saqa_id,
                                                 'learning programme id': learner.lp_learner_assessment_line_id.code,
                                                 'skills programme id': '',
                                                 })

                if assessment.qual_skill_assessment == 'skill':
                    for learner in assessment.assessment.learner_ids_for_skills:
                        lnr_count_sp = len(list(set(assessment.assessment.learner_ids_for_skills)))

                        sa_citizen = False
                        foreign = False
                        lnr = learner.learner_id

                        if lnr.alternate_id_type:
                            foreign = True

                        # special case where both fields are filled
                        if lnr.learner_identification_id and lnr.national_id:
                            foreign = False

                        lnr.person_last_name = remove_ascii(lnr.person_last_name)
                        lnr.name = remove_ascii(lnr.name)

                        for quals in lnr.skills_programme_ids:
                            if assessment.batch_id == quals.batch_id:
                                writer.writerow({'NAME': '',
                                                 'provider': '',
                                                 'type': '',
                                                 'batch': '',
                                                 'state': '',
                                                 'province': '',
                                                 'enrolled learners': '',
                                                 'learner ID': lnr.learner_identification_id,
                                                 'foreign ID': lnr.national_id if foreign else '',
                                                 'first name': lnr.name,
                                                 'last name': lnr.person_last_name,
                                                 'employed': lnr.socio_economic_status,
                                                 'achieved': quals.is_learner_achieved,
                                                 'qualification': learner.skill_learner_assessment_line_id.name,
                                                 'qualification id': quals.saqa_skill_id,
                                                 'learning programme id': '',
                                                 'skills programme id': learner.skill_learner_assessment_line_id.code,
                                                 })

        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=assessment_analysis.csv;')],
                                         cookies={})
        # dbg(response)
        # import os
        # dbg(os.getcwd())
        with open('/odoo_reports/assessment_analysis.csv', 'r') as f2:
            data = str.encode(f2.read(), 'utf-8')
            response.response = data

        return response

    @http.route(['/report_export/assessment_approval_analysis/<int:report_id>'], type='http', auth="user")
    def assessment_approval_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        assessments = request.env['seta.reports.assessment.approval'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        #  FROM %s TO %s") % (report.from_date, report.to_date)
        worksheet.write_merge(0, 0, 0, 8, _("Learner Achievement recommendations from the Provinces Approved_Declined by "
                                            "ETQA Head Office FROM %s TO %s") % (report.from_date, report.to_date),
                              header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, assessment in enumerate(assessments):
            in_progress = assessment.total - (assessment.approved_count + assessment.denied_count)
            worksheet.write(i + 2, 0, assessment.province.name)
            worksheet.write(i + 2, 1, assessment.total)
            worksheet.write(i + 2, 2, assessment.approved_count)
            # worksheet.write(i + 2, 3, assessment.denied_count)        #TODO ASK why this was commented
            worksheet.write(i + 2, 4, assessment.approved_perc)
            # worksheet.write(i + 2, 5, assessment.denied_perc)         # As well as this
            worksheet.write(i + 2, 6, in_progress)

        # num_leads = len(leads)
        # worksheet.write_merge(num_leads + 2, num_leads + 2, 22, 23, "TOTAL", header_bold_blue)
        # worksheet.write(num_leads + 2, 24, xlwt.Formula("SUM($X$3:$X$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 25, xlwt.Formula("SUM($Y$3:$Y$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 26, xlwt.Formula("SUM($Z$3:$Z$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 27, xlwt.Formula("SUM($AA$3:$AA$%s)" % (num_leads + 2)), header_bold_yellow)
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=assessment_approval_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    @http.route(['/report_export/accreditation_etqa_approval_analysis/<int:report_id>'], type='http', auth="user")
    def accreditation_etqa_approval_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        assessments = request.env['seta.reports.etqa.approval.accreditation.analysis'].search(
            [('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        # report FROM %s TO %s") % (report.from_date, report.to_date)
        worksheet.write_merge(0, 0, 0, 8, _("Provider Accreditation recommendation from Provinces_Approved_Declined"
                              " FROM %s TO %s") % (report.from_date, report.to_date), header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, assessment in enumerate(assessments):
            in_progress = assessment.total - (assessment.approved_count + assessment.denied_count)
            worksheet.write(i + 2, 0, assessment.province.name)
            worksheet.write(i + 2, 1, assessment.total)
            worksheet.write(i + 2, 2, assessment.approved_count)
            worksheet.write(i + 2, 3, assessment.denied_count)
            worksheet.write(i + 2, 4, assessment.approved_perc)
            worksheet.write(i + 2, 5, assessment.denied_perc)
            worksheet.write(i + 2, 6, in_progress)
            worksheet.write(i + 2, 7, assessment.new_accred)
            worksheet.write(i + 2, 8, assessment.new_prog_approval)
            worksheet.write(i + 2, 9, assessment.reaccred)
            worksheet.write(i + 2, 10, assessment.extension)


        # num_leads = len(leads)
        # worksheet.write_merge(num_leads + 2, num_leads + 2, 22, 23, "TOTAL", header_bold_blue)
        # worksheet.write(num_leads + 2, 24, xlwt.Formula("SUM($X$3:$X$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 25, xlwt.Formula("SUM($Y$3:$Y$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 26, xlwt.Formula("SUM($Z$3:$Z$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 27, xlwt.Formula("SUM($AA$3:$AA$%s)" % (num_leads + 2)), header_bold_yellow)
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=accreditation_etqa_approval_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    @http.route(['/report_export/register_approval_analysis/<int:report_id>'], type='http', auth="user")
    def register_approval_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        registrations = request.env['seta.reports.register'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        # FROM %s TO %s") % (report.from_date, report.to_date)
        worksheet.write_merge(0, 0, 0, 7, _("%% of Approvals and declines of %s per applications per province FROM %s TO %s") %
                             (report.register_assessor_or_moderator, report.from_date, report.to_date), header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, reg in enumerate(registrations):
            dbg('########  ' + str(reg))
            in_progress = reg.total - (reg.approved_count + reg.denied_count)
            worksheet.write(i + 2, 0, reg.province.name)
            worksheet.write(i + 2, 1, reg.total)
            worksheet.write(i + 2, 2, reg.approved_count)
            worksheet.write(i + 2, 3, reg.denied_count)
            worksheet.write(i + 2, 4, reg.approved_perc)
            worksheet.write(i + 2, 5, reg.denied_perc)
            worksheet.write(i + 2, 6, in_progress)
            worksheet.write(i + 2, 7, reg.new_registration_count)
            worksheet.write(i + 2, 8, reg.re_registration_count)
            worksheet.write(i + 2, 9, reg.extension_of_scope_count)



        # num_leads = len(leads)
        # worksheet.write_merge(num_leads + 2, num_leads + 2, 22, 23, "TOTAL", header_bold_blue)
        # worksheet.write(num_leads + 2, 24, xlwt.Formula("SUM($X$3:$X$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 25, xlwt.Formula("SUM($Y$3:$Y$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 26, xlwt.Formula("SUM($Z$3:$Z$%s)" % (num_leads + 2)), header_bold_yellow)
        # worksheet.write(num_leads + 2, 27, xlwt.Formula("SUM($AA$3:$AA$%s)" % (num_leads + 2)), header_bold_yellow)
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=register_approval_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    @http.route(['/report_export/mod_ass_register_8week_analysis/<int:report_id>'], type='http', auth="user")
    def mod_ass_register_8week_analysis(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        registrations = request.env['seta.reports.8week.register'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(report[0].name)
        # worksheet = workbook.add_worksheet(report[0].name)
        header_bold_blue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
        header_bold_lightblue = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
        header_bold_yellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_bold_lightyellow = xlwt.easyxf(
            "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
        header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
        bold = xlwt.easyxf("font: bold on;")
        normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
        # Step 1: writing headers
        # FROM %s TO %s") % (report.from_date, report.to_date)
        worksheet.write_merge(0, 0, 0, 9, _("%s applications approved_declined  within the 8_week period_the "
                                            "whole process per province FROM %s TO %s") % (report.register_assessor_or_moderator, report.from_date, report.to_date),
                              header_bold_blue)

        for i, header in enumerate(headers):
            worksheet.write(1, i, header, header_bold_lightblue)

        for i, reg in enumerate(registrations):
            worksheet.write(i + 2, 0, reg.mod_id_no)
            worksheet.write(i + 2, 1, reg.mod_name)
            worksheet.write(i + 2, 2, reg.mod_surname)
            worksheet.write(i + 2, 3, reg.province.name)
            worksheet.write(i + 2, 4, reg.application_date)
            worksheet.write(i + 2, 5, reg.update_date)
            worksheet.write(i + 2, 6, reg.days_to_update)
            worksheet.write(i + 2, 7, reg.final_state)
            worksheet.write(i + 2, 8, reg.in_process)
            worksheet.write(i + 2, 9, reg.transaction_type)

        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=mod_ass_register_8week_analysis.xls;')],
                                         cookies={})
        workbook.save(response.stream)

        return response

    # @http.route(['/report_export/sdps_no_learners/<int:report_id>'], type='http', auth="user")
    # def sdps_no_learners(self, report_id, **kw):
    #     # jdata = json.loads(data)
    #
    #     report = request.env['seta.reports'].search([('id', '=', report_id)])
    #     providers = request.env['seta.reports.etqa.sdps.no.learners'].search([('report_id', '=', report_id)])
    #     headers = ast.literal_eval(report.headers)
    #
    #     workbook = xlwt.Workbook()
    #     worksheet = workbook.add_sheet(report[0].name)
    #     # worksheet = workbook.add_worksheet(report[0].name)
    #     header_bold_blue = xlwt.easyxf(
    #         "font: bold on; pattern: pattern solid, fore_colour blue; align: vert center, horiz center;")
    #     header_bold_lightblue = xlwt.easyxf(
    #         "font: bold on; pattern: pattern solid, fore_colour blue; align: horiz center;")
    #     header_bold_yellow = xlwt.easyxf(
    #         "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
    #     header_bold_lightyellow = xlwt.easyxf(
    #         "font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center;")
    #     header_plain = xlwt.easyxf("pattern: pattern solid, fore_colour blue;")
    #     bold = xlwt.easyxf("font: bold on;")
    #     normal_yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow; align: horiz right;")
    #     # Step 1: writing headers
    #     worksheet.write_merge(0, 0, 0, 5, _("registrations report FROM %s TO %s") % (report.from_date, report.to_date),
    #                           header_bold_blue)
    #
    #     for i, header in enumerate(headers):
    #         worksheet.write(1, i, header, header_bold_lightblue)
    #     #[_('NAME'), _('Provider Accreditation Number'), ('Primary Accrediting Body'),
    #     # ('Accreditation Start Date'), ('Accreditation Start Date'), ('Email Address'),
    #     # ('Physical Address'), ('Province'), ('Accredited Qualification Title'),
    #     # ('Qualification ID')]
    #
    #     # # for provider in providers:
    #     # #     if provider.qualification_ids:
    #     # #         worksheet.write()
    #     # #         for qualification in provider.qualification_ids:
    #     #
    #     #     if provider.
    #     master = 2
    #     for i, provider in enumerate(providers):
    #         dbg('master i ')
    #         dbg(master)
    #         dbg(provider.provider_id.name)
    #         if provider.provider_id.physical_address_1 and provider.provider_id.physical_address_2:
    #             addr = provider.provider_id.physical_address_1 or '' + ',' + provider.provider_id.physical_address_2 or ''
    #         else:
    #             addr = 'problems!!!!!!!!!'
    #         # worksheet.write(i + 2, 0, provider.provider_id.name)
    #         # worksheet.write(i + 2, 1, provider.provider_id.provider_accreditation_num)
    #         # worksheet.write(i + 2, 2, 'accred body?')
    #         # worksheet.write(i + 2, 3, provider.provider_id.provider_start_date)
    #         # worksheet.write(i + 2, 4, provider.provider_id.provider_end_date)
    #         # worksheet.write(i + 2, 5, provider.provider_id.email)
    #         # worksheet.write(i + 2, 6, addr)
    #         # worksheet.write(i + 2, 7, provider.provider_id.province_code_physical.name)
    #
    #         worksheet.write(master, 0, provider.provider_id.name)
    #         worksheet.write(master, 1, provider.provider_id.provider_accreditation_num)
    #         worksheet.write(master, 2, 'accred body?')
    #         worksheet.write(master, 3, provider.provider_id.provider_start_date)
    #         worksheet.write(master, 4, provider.provider_id.provider_end_date)
    #         worksheet.write(master, 5, provider.provider_id.email)
    #         worksheet.write(master, 6, addr)
    #         worksheet.write(master, 7, provider.provider_id.province_code_physical.name)
    #         master += i
    #         for j, qual in enumerate(provider.qualification_ids):
    #             master += j
    #             dbg('master j')
    #             dbg(master)
    #             dbg(qual.saqa_qual_id)
    #             worksheet.write(master, 8, qual.qualification_id.name)
    #             worksheet.write(master, 9, qual.saqa_qual_id)
    #         # for qual in provider.qualification_ids:
    #         #     worksheet.write(i + 2, 8, qual.qualification_id.name)
    #         #     worksheet.write(i + 2, 9, qual.saqa_qual_id)
    #         # for skill in provider.skill_ids:
    #         #     worksheet.write(i + 2, 10, skill.skills_programme_id.name)
    #         #     worksheet.write(i + 2, 11, skill.skill_saqa_id)
    #         # for lp in provider.lp_ids:
    #         #     worksheet.write(i + 2, 12, lp.learning_programme_id.name)
    #         #     worksheet.write(i + 2, 13, lp.lp_saqa_id)
    #
    #     response = request.make_response(None,
    #                                      headers=[('Content-Type', 'application/vnd.ms-excel'),
    #                                               ('Content-Disposition',
    #                                                'attachment; filename=sdps_no_learners.xls;')],
    #                                      cookies={})
    #     workbook.save(response.stream)
    #
    #     return response

    @http.route(['/report_export/sdps_no_learners/<int:report_id>'], type='http', auth="user")
    def sdps_no_learners(self, report_id, **kw):
        # jdata = json.loads(data)

        report = request.env['seta.reports'].search([('id', '=', report_id)])
        providers = request.env['seta.reports.etqa.sdps.no.learners'].search([('report_id', '=', report_id)])
        headers = ast.literal_eval(report.headers)
        # using csv to avoid row limit and nesting crap(dont go back to trying xlsx for long list reports/master data lopps)
        with open('/odoo_reports/sdps_no_learners.csv', 'w') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers, delimiter=',', quotechar='"')

            writer.writeheader()
            for provider in providers:
                dbg('writing provider to file:' + str(provider.provider_id.name))
                # below used as lazy way to catch concat issues/side effect :may trigger data correction
                if provider.provider_id.physical_address_1 and provider.provider_id.physical_address_2:
                    addr = provider.provider_id.physical_address_1 or '' + ',' + provider.provider_id.physical_address_2 or ''
                else:
                    addr = 'Data Issue'
                start = provider.start_date
                end = provider.end_date
                if provider.provider_id.optYesNo:
                    seta_opt = 'Other Seta'
                else:
                    seta_opt = 'HWSeta'
                writer.writerow({'NAME': provider.provider_id.name,
                                 'Provider Accreditation Number': provider.provider_id.provider_accreditation_num,
                                 'Primary Accrediting Body': seta_opt,
                                 'Accreditation Start Date': provider.provider_id.provider_start_date,
                                 'Accreditation End Date': provider.provider_id.provider_end_date,
                                 'Email Address': provider.provider_id.email,
                                 'Physical Address': addr,
                                 'Province': provider.provider_id.province_code_physical.name,
                                 'Type': '',
                                 'Accredited Qualification Title': '',
                                 'Qualification ID': '',
                                 'Skill ID': '',
                                 'LP ID': '',
                                 'Learners Enrolled': '',
                                 })
                for qual in provider.qualification_ids:
                    # todo: remove or buff to avoid misleading stats
                    learners = request.env['hr.employee'].search(
                        [('learner_qualification_ids.provider_id', '=', provider.provider_id.id),
                         ('learner_qualification_ids.learner_qualification_parent_id', '=', qual.qualification_id.id)])
                    # gives incorrect total at first glance
                    # learner_reg = request.env['learner.registration'].search([('learner_qualification_ids.create_date','>',start),('learner_qualification_ids.create_date','<',end),('learner_qualification_ids.provider_id','=',provider.provider_id.id),('learner_qualification_ids.learner_qualification_parent_id','=',qual.qualification_id.id)])
                    # dont use create date, seems to destroy expectations from imports
                    learner_reg_lines = request.env['learner.registration.qualification'].search(
                        [('start_date', '>', start), ('start_date', '<', end),
                         ('provider_id', '=', provider.provider_id.id),
                         ('learner_qualification_parent_id', '=', qual.qualification_id.id)])
                    writer.writerow({'NAME': '',
                                     'Provider Accreditation Number': '',
                                     'Primary Accrediting Body': '',
                                     'Accreditation Start Date': '',
                                     'Accreditation End Date': '',
                                     'Email Address': '',
                                     'Physical Address': '',
                                     'Province': '',
                                     'Type': 'Qualification',
                                     'Accredited Qualification Title': qual.qualification_id.name,
                                     'Qualification ID': qual.saqa_qual_id,
                                     'Skill ID': '',
                                     'LP ID': '',
                                     'Learners Enrolled': len(learner_reg_lines),
                                     'Learners Total': len(learners),  # this needs to be removed,as for all types
                                     })
                for skill in provider.skill_ids:
                    learners = request.env['hr.employee'].search(
                        [('skills_programme_ids.provider_id', '=', provider.provider_id.id),
                         ('skills_programme_ids.skills_programme_id', '=', skill.id)])
                    # learner_reg = request.env['hr.employee'].search(
                    #     [('skills_programme_ids.create_date','>',start),('skills_programme_ids.create_date','<',end),
                    #      ('skills_programme_ids.provider_id', '=', provider.provider_id.id),
                    #      ('skills_programme_ids.skills_programme_id', '=', skill.id)])
                    # dont use create date, seems to destroy expectations from imports
                    learner_reg_lines = request.env['skills.programme.learner.rel'].search(
                        [('start_date', '>', start), ('start_date', '<', end),
                         ('provider_id', '=', provider.provider_id.id),
                         ('skills_programme_id', '=', skill.id)])
                    writer.writerow({'NAME': '',
                                     'Provider Accreditation Number': '',
                                     'Primary Accrediting Body': '',
                                     'Accreditation Start Date': '',
                                     'Accreditation End Date': '',
                                     'Email Address': '',
                                     'Physical Address': '',
                                     'Province': '',
                                     'Type': 'Skill',
                                     'Accredited Qualification Title': skill.skills_programme_id.name,
                                     'Qualification ID': skill.skill_saqa_id,
                                     'Skill ID': skill.skills_programme_id.code,
                                     'LP ID': '',
                                     'Learners Enrolled': len(learner_reg_lines),
                                     'Learners Total': len(learners),
                                     })
                for lp in provider.lp_ids:
                    learners = request.env['hr.employee'].search(
                        [('learning_programme_ids.provider_id', '=', provider.provider_id.id),
                         ('learning_programme_ids.learning_programme_id', '=', lp.id)])
                    # learner_reg = request.env['hr.employee'].search(
                    #     [('learning_programme_ids.create_date','>',start),('learning_programme_ids.create_date','<',end),
                    #      ('learning_programme_ids.provider_id', '=', provider.provider_id.id),
                    #      ('learning_programme_ids.learning_programme_id', '=', lp.id)])
                    # dont use create date, seems to destroy expectations from imports
                    learner_reg_lines = request.env['learning.programme.learner.rel'].search(
                        [('start_date', '>', start), ('start_date', '<', end),
                         ('provider_id', '=', provider.provider_id.id),
                         ('learning_programme_id', '=', lp.id)])
                    writer.writerow({'NAME': '',
                                     'Provider Accreditation Number': '',
                                     'Primary Accrediting Body': '',
                                     'Accreditation Start Date': '',
                                     'Accreditation End Date': '',
                                     'Email Address': '',
                                     'Physical Address': '',
                                     'Province': '',
                                     'Type': 'LP',
                                     'Accredited Qualification Title': lp.learning_programme_id.name,
                                     'Qualification ID': lp.lp_saqa_id,
                                     'Skill ID': '',
                                     'LP ID': lp.learning_programme_id.code,
                                     'Learners Enrolled': len(learner_reg_lines),
                                     'Learners Total': len(learners),
                                     })
        response = request.make_response(None,
                                         headers=[('Content-Type', 'application/vnd.ms-excel'),
                                                  ('Content-Disposition',
                                                   'attachment; filename=sdps_no_learners.csv;')],
                                         cookies={})
        dbg(response)
        import os
        dbg(os.getcwd())
        with open('/odoo_reports/sdps_no_learners.csv', 'r') as f2:
            data = str.encode(f2.read(), 'utf-8')
            response.response = data

        return response
